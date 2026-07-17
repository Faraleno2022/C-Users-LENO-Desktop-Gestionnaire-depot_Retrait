"""Génération de rapports côté serveur (Excel + PDF) à partir des modèles Django.

Réplique fonctionnelle de `app/services/report_service.py` (côté poste) et
`app/utils/exporters.py`, adapté pour streamer les fichiers en réponse HTTP.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from sync.models import Product, Sale, StockMovement, Transaction


# --- Périodes ----------------------------------------------------------------

def today_period() -> Tuple[str, str]:
    t = date.today().strftime("%Y-%m-%d")
    return t, t


def month_period() -> Tuple[str, str]:
    today = date.today()
    start = today.replace(day=1)
    if today.month == 12:
        nxt = today.replace(year=today.year + 1, month=1, day=1)
    else:
        nxt = today.replace(month=today.month + 1, day=1)
    end = nxt - timedelta(days=1)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def year_period() -> Tuple[str, str]:
    y = date.today().year
    return f"{y}-01-01", f"{y}-12-31"


# --- Formatters --------------------------------------------------------------

def _fmt_money(v) -> str:
    try:
        return f"{float(v):,.0f} GNF".replace(",", " ")
    except (TypeError, ValueError):
        return "0 GNF"


def _fmt_num(v) -> str:
    try:
        return f"{float(v):g}"
    except (TypeError, ValueError):
        return str(v or "")


# --- Récupération des données -------------------------------------------------

TRANSACTION_HEADERS = [
    "ID", "Date", "Matricule", "Téléphone", "Type",
    "Dépôt", "Retrait", "Solde après", "Agent",
]

SALES_HEADERS = [
    "ID", "Date", "Matricule", "Produit", "Quantité",
    "Prix unitaire", "Montant", "Solde après", "Agent",
]

STOCK_HEADERS = [
    "ID", "Référence", "Produit", "Prix unitaire", "Stock",
    "Seuil", "Valeur", "Statut",
]

STOCK_MOVEMENTS_HEADERS = [
    "Date", "Produit", "Type", "Quantité", "Stock après", "Motif", "Agent",
]


def fetch_stock_movements(date_from: str, date_to: str, agent: Optional[str] = None) -> List[Sequence]:
    """Mouvements de stock (entrées/sorties) sur une période — utilisable par
    mois ou par année via les périodes du formulaire de rapport."""
    qs = StockMovement.objects.filter(
        deleted=False,
        created_at__gte=f"{date_from} 00:00:00",
        created_at__lte=f"{date_to} 23:59:59",
    )
    if agent:
        qs = qs.filter(agent_nom__icontains=agent)
    qs = qs.order_by("created_at", "id")
    rows = []
    for m in qs:
        rows.append([
            m.created_at,
            m.product_nom,
            "Entrée" if m.type == "entree" else "Sortie",
            _fmt_num(m.quantite),
            _fmt_num(m.stock_apres),
            m.motif or "",
            m.agent_nom or "",
        ])
    return rows


def fetch_transactions(date_from: str, date_to: str, agent: Optional[str] = None) -> List[Sequence]:
    qs = Transaction.objects.filter(
        deleted=False,
        created_at__gte=f"{date_from} 00:00:00",
        created_at__lte=f"{date_to} 23:59:59",
    )
    if agent:
        qs = qs.filter(agent_nom__icontains=agent)
    qs = qs.order_by("created_at", "id")
    rows = []
    for t in qs:
        rows.append([
            t.id,
            t.created_at,
            t.matricule,
            t.telephone or "",
            "Dépôt" if t.type == "depot" else "Retrait",
            _fmt_money(t.montant) if t.type == "depot" else "",
            _fmt_money(t.montant) if t.type == "retrait" else "",
            _fmt_money(t.solde_apres),
            t.agent_nom,
        ])
    return rows


def fetch_sales(date_from: str, date_to: str, agent: Optional[str] = None) -> List[Sequence]:
    qs = Sale.objects.filter(
        deleted=False,
        created_at__gte=f"{date_from} 00:00:00",
        created_at__lte=f"{date_to} 23:59:59",
    )
    if agent:
        qs = qs.filter(agent_nom__icontains=agent)
    qs = qs.order_by("created_at", "id")
    rows = []
    for s in qs:
        rows.append([
            s.id,
            s.created_at,
            s.matricule,
            s.product_nom,
            _fmt_num(s.quantite),
            _fmt_money(s.prix_unitaire),
            _fmt_money(s.montant_total),
            _fmt_money(s.solde_apres),
            s.agent_nom,
        ])
    return rows


def fetch_stock() -> List[Sequence]:
    rows = []
    for p in Product.objects.all().order_by("nom"):
        if not p.actif:
            statut = "Inactif"
        elif (p.quantite_stock or 0) <= (p.seuil_alerte or 0):
            statut = "Stock bas"
        else:
            statut = "OK"
        valeur = (p.quantite_stock or 0) * (p.prix_unitaire or 0)
        rows.append([
            p.id,
            p.reference or "",
            p.nom,
            _fmt_money(p.prix_unitaire),
            _fmt_num(p.quantite_stock),
            _fmt_num(p.seuil_alerte),
            _fmt_money(valeur),
            statut,
        ])
    return rows


# --- Génération de fichier en mémoire ----------------------------------------

def _safe_sheet_title(title: str) -> str:
    """Nom d'onglet Excel valide : Excel interdit / \\ ? * [ ] : et limite à 31 car.

    Sans ce nettoyage, un titre comme « Dépôts / Retraits » fait planter l'export
    (ValueError: Invalid character / found in sheet title).
    """
    cleaned = title or "Export"
    for ch in "/\\?*[]:":
        cleaned = cleaned.replace(ch, "-")
    cleaned = cleaned.strip() or "Export"
    return cleaned[:31]


def build_excel(title: str, headers: Sequence[str], rows: Sequence[Sequence]) -> bytes:
    wb = Workbook()
    ws = wb.active
    # Le nom d'onglet est nettoyé ; la cellule A1 garde le titre complet (les
    # cellules acceptent tous les caractères, contrairement aux noms d'onglets).
    ws.title = _safe_sheet_title(title)
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(list(headers))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(list(row))
    for col_idx, _ in enumerate(headers, start=1):
        max_len = 12
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True).__next__():
            if cell is None:
                continue
            max_len = max(max_len, min(40, len(str(cell)) + 2))
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max_len
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(title: str, headers: Sequence[str], rows: Sequence[Sequence], subtitle: str = "") -> bytes:
    buf = io.BytesIO()
    page_size = landscape(A4) if len(headers) > 6 else A4
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(f"<b>{title}</b>", styles["Title"])]
    if subtitle:
        story.append(Paragraph(subtitle, styles["Normal"]))
    story.append(Spacer(1, 6 * mm))

    data = [list(headers)] + [list(map(_pdf_cell, row)) for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def _pdf_cell(value) -> str:
    if value is None:
        return ""
    return str(value)


def build_withdrawal_invoice(data: dict) -> bytes:
    """Génère une facture PDF pour un retrait / vente de produits.

    `data` correspond à la structure stockée dans la session (`last_withdrawal`) :
    id, matricule, telephone, agent_nom, created_at, solde_avant, montant,
    solde_apres, lines[{product_nom, quantite, prix, total}].
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    story = []

    # En-tête société
    story.append(Paragraph("<b>EMAB GROUP</b>", styles["Title"]))
    story.append(Paragraph("Facture de retrait / vente de produits", styles["Heading2"]))
    story.append(Spacer(1, 4 * mm))

    # Métadonnées facture / client
    meta = [
        [f"Facture N° : {data.get('id', '')}", f"Date : {data.get('created_at', '')}"],
        [f"Matricule : {data.get('matricule', '')}", f"Téléphone : {data.get('telephone') or '—'}"],
        [f"Agent : {data.get('agent_nom') or '—'}", ""],
    ]
    mt = Table(meta, colWidths=[doc.width / 2.0, doc.width / 2.0])
    mt.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(mt)
    story.append(Spacer(1, 5 * mm))

    # Lignes de produits
    lines = data.get("lines") or []
    if lines:
        tbl = [["Produit", "Qté", "Prix unitaire", "Sous-total"]]
        for l in lines:
            tbl.append([
                _pdf_cell(l.get("product_nom")),
                _fmt_num(l.get("quantite")),
                _fmt_money(l.get("prix")),
                _fmt_money(l.get("total")),
            ])
        t = Table(
            tbl,
            colWidths=[doc.width * 0.46, doc.width * 0.14, doc.width * 0.20, doc.width * 0.20],
            repeatRows=1,
        )
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]))
        story.append(t)
        story.append(Spacer(1, 4 * mm))

    # Récapitulatif des soldes
    recap = [
        ["Solde avant", _fmt_money(data.get("solde_avant"))],
        ["Montant du retrait", "- " + _fmt_money(data.get("montant"))],
        ["Solde restant", _fmt_money(data.get("solde_apres"))],
    ]
    rt = Table(recap, colWidths=[doc.width * 0.6, doc.width * 0.4])
    rt.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#b00020")),
        ("LINEABOVE", (0, 0), (-1, 0), 0.5, colors.grey),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(rt)
    story.append(Spacer(1, 12 * mm))
    story.append(Paragraph("Merci de votre confiance — EMAB GROUP", styles["Italic"]))

    doc.build(story)
    return buf.getvalue()


# --- Façade : dataset + kind + format -> (filename, mime, bytes) -------------

KIND_TITLES = {
    "daily": "Rapport journalier",
    "monthly": "Rapport mensuel",
    "annual": "Rapport annuel",
    "agent": "Rapport par agent",
    "custom": "Rapport personnalisé",
    "all": "Tout l'historique",
}

DATASET_TITLES = {
    "transactions": "Dépôts / Retraits",
    "sales": "Ventes",
    "stock": "État du stock",
    "stock_movements": "Mouvements de stock",
}


def generate(
    dataset: str, kind: str, date_from: str, date_to: str,
    agent: Optional[str] = None, fmt: str = "pdf",
) -> Tuple[str, str, bytes]:
    """Renvoie (filename, mime_type, content_bytes)."""
    if dataset == "sales":
        rows = fetch_sales(date_from, date_to, agent=agent)
        headers = SALES_HEADERS
        title = f"{DATASET_TITLES['sales']}"
    elif dataset == "stock":
        rows = fetch_stock()
        headers = STOCK_HEADERS
        title = DATASET_TITLES["stock"]
    elif dataset == "stock_movements":
        rows = fetch_stock_movements(date_from, date_to, agent=agent)
        headers = STOCK_MOVEMENTS_HEADERS
        title = DATASET_TITLES["stock_movements"]
    else:
        dataset = "transactions"
        rows = fetch_transactions(date_from, date_to, agent=agent)
        headers = TRANSACTION_HEADERS
        title = DATASET_TITLES["transactions"]

    kind_label = KIND_TITLES.get(kind, "Rapport personnalisé")
    if dataset == "stock":
        # « État du stock » = photo actuelle (sans période).
        subtitle = f"Au {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    else:
        subtitle = f"Du {date_from} au {date_to}"
        if agent:
            subtitle += f" — Agent : {agent}"
    full_title = f"{title} — {kind_label}"

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = f"{dataset}_{kind}"
    if fmt == "xlsx":
        content = build_excel(f"{full_title} ({subtitle})", headers, rows)
        return f"rapport_{safe}_{ts}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", content
    content = build_pdf(full_title, headers, rows, subtitle=subtitle)
    return f"rapport_{safe}_{ts}.pdf", "application/pdf", content
