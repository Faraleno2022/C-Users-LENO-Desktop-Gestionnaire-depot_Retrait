"""Export Excel et PDF."""
from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


def export_to_excel(file_path: Path, title: str, headers: Sequence[str], rows: Sequence[Sequence]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30] or "Export"
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append([])

    ws.append(list(headers))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(list(row))

    for col_idx, _ in enumerate(headers, start=1):
        max_len = 12
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True).__next__():
            if cell is None:
                continue
            max_len = max(max_len, min(40, len(str(cell)) + 2))
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max_len

    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(file_path))
    return file_path


def export_to_pdf(file_path: Path, title: str, headers: Sequence[str], rows: Sequence[Sequence], subtitle: str = "") -> Path:
    file_path.parent.mkdir(parents=True, exist_ok=True)
    page_size = landscape(A4) if len(headers) > 6 else A4
    doc = SimpleDocTemplate(
        str(file_path),
        pagesize=page_size,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(f"<b>{title}</b>", styles["Title"])]
    if subtitle:
        story.append(Paragraph(subtitle, styles["Normal"]))
    story.append(Spacer(1, 6 * mm))

    data = [list(headers)] + [list(map(_pdf_cell, row)) for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return file_path


def _pdf_cell(value) -> str:
    if value is None:
        return ""
    return str(value)
